#!/usr/bin/env python3
"""
Convert NJ Portal Search Excel file to CSV
Handles Excel files with formatting issues by using openpyxl directly
"""

import openpyxl
import csv
from pathlib import Path

EXCEL_FILE = Path(__file__).parent.parent / "output" / "state_licenses" / "new_jersey" / "portal_search_nj.xlsx"
CSV_FILE = Path(__file__).parent.parent / "output" / "state_licenses" / "new_jersey" / "portal_search_nj.csv"

print(f"Reading Excel file: {EXCEL_FILE}")
print(f"File size: {EXCEL_FILE.stat().st_size:,} bytes")

# Use read_only mode to bypass styling issues
print("\nUsing openpyxl in read_only mode (bypasses styling issues)...")
workbook = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)
sheet = workbook.active
print(f"✅ Successfully loaded workbook!")
print(f"Active sheet: {workbook.sheetnames[0] if workbook.sheetnames else 'Unknown'}")

# Extract all data
rows = []
for row in sheet.iter_rows(values_only=True):
    rows.append(row)

print(f"Extracted {len(rows):,} rows")

# First row is header
if rows:
    header = rows[0]
    data_rows = rows[1:]

    print(f"\nColumns ({len(header)}):")
    for i, col in enumerate(header, 1):
        print(f"  {i}. {col}")

    # Save to CSV
    with open(CSV_FILE, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(header)
        writer.writerows(data_rows)

    print(f"\n✅ Saved {len(data_rows):,} contractors to: {CSV_FILE}")

    # Show preview
    print(f"\n{'='*80}")
    print("PREVIEW - First 5 rows:")
    print('='*80)
    for i, row in enumerate(data_rows[:5], 1):
        print(f"\n  Row {i}:")
        for col_name, value in zip(header, row):
            if value:  # Only show non-empty values
                print(f"    {col_name}: {value}")

    # Show statistics - count by license type if column exists
    print(f"\n{'='*80}")
    print("STATISTICS:")
    print('='*80)

    # Find license type column
    type_col_idx = None
    for i, col in enumerate(header):
        if col and any(word in str(col).lower() for word in ['license', 'type', 'profession', 'category']):
            type_col_idx = i
            print(f"\nLicense type column: {col}")
            break

    if type_col_idx is not None:
        # Count license types
        type_counts = {}
        for row in data_rows:
            if len(row) > type_col_idx:
                license_type = row[type_col_idx]
                if license_type:
                    type_counts[license_type] = type_counts.get(license_type, 0) + 1

        print(f"\nLicense type distribution:")
        for license_type, count in sorted(type_counts.items(), key=lambda x: x[1], reverse=True):
            print(f"  {license_type}: {count:,}")
        print(f"\nTotal unique license types: {len(type_counts)}")

    print(f"\nTotal contractors: {len(data_rows):,}")
